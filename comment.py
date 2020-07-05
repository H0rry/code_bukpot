from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font

folder_png = "results"
excel_name = "Incubator_Bukpot_Garry.xlsx"

wb = load_workbook(excel_name)
ws = wb.active

row_count = ws.max_row


def comment_and_list_duplicate():
    for sheet in wb.worksheets:

        list_address_kosong = []
        dk = {}
        tbdel = []

        for row1 in range(2, row_count + 1):
            for row2 in range(2, row_count + 1):
                if row1 == row2:
                    continue
                cell1_npwp = sheet[row1][3]
                cell2_npwp = sheet[row2][3]
                cell1_bukpot = sheet[row1][4]
                cell2_bukpot = sheet[row2][4]
                cell_file1 = sheet[row1][1]
                cell_file2 = sheet[row2][1]
                if (
                    cell1_bukpot.value == cell2_bukpot.value
                    and cell1_npwp.value == cell2_npwp.value
                    and f"{cell1_npwp}{cell2_npwp}" not in list_address_kosong
                ):
                    list_address_kosong.append(f"{cell1_npwp}{cell2_npwp}")
                    list_address_kosong.append(f"{cell2_npwp}{cell1_npwp}")
                    give_comment(sheet.title, cell1_bukpot, cell2_bukpot)

                    if cell_file1 not in tbdel:
                        if cell_file1 not in dk.keys():
                            dk[cell_file1] = []
                        dk[cell_file1].append(cell_file2)
                        tbdel.append(cell_file2)

        make_list_duplicate(dk, sheet)

    wb.save(excel_name)


def give_comment(sheet_title, cell1, cell2):
    address1 = cell1.row - 1
    address2 = cell2.row - 1
    print(f"give comment at (sheet {sheet_title}) No {address1} and {address2}")

    comment1 = Comment(f">> No = {address2}", "")
    comment2 = Comment(f">> No = {address1}", "")

    if cell1.comment:
        comment1 = Comment(f"{cell1.comment.text}, {address2}", "")

    if cell2.comment:
        comment2 = Comment(f"{cell2.comment.text}, {address1}", "")

    cell1.comment = comment1
    cell2.comment = comment2


def make_list_duplicate(dk, sheet):
    title_keep_cell = sheet[row_count + 2][1]
    title_remove_cell = sheet[row_count + 2][2]

    title_keep_cell.value = "list_keep_file"
    title_remove_cell.value = "list_remove_file"

    isi_keep = title_keep_cell.offset(row=1)
    isi_remove = title_remove_cell.offset(row=1)

    for keep_file, remove_list in dk.items():
        isi_keep.value = keep_file.value
        isi_keep.hyperlink = f"{folder_png}/{keep_file.value}.png"
        isi_keep.font = Font(color="000000FF")
        isi_keep = isi_keep.offset(row=1)
        for remove_file in remove_list:
            isi_remove.value = remove_file.value
            isi_remove.hyperlink = f"{folder_png}/{remove_file.value}.png"
            isi_remove.font = Font(color="000000FF")
            isi_remove = isi_remove.offset(column=1)
        isi_remove = isi_keep.offset(column=1)


comment_and_list_duplicate()
