import os
import re

import pandas as pd
import progressbar
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, NamedStyle
from openpyxl.formatting.rule import CellIsRule, FormulaRule


import ocr_png


def make_list_of_files(folder_png):
    files = []

    for file in os.listdir(folder_png):
        if file.endswith(".png"):
            files.append(os.path.join(file))

    files.sort(key=natural_keys)
    return files


def natural_keys(text):  # natural sorting
    return [atoi(c) for c in re.split(r"(\d+)", text)]


def atoi(text):
    return int(text) if text.isdigit() else text


def make_dict_of_dfs(files):
    dict_of_dfs = {}

    print(f"\nTotal Bukti Potong = {len(files)} halaman\n")
    print("Starting OCR:")

    bar = progressbar.ProgressBar(
        maxval=len(files),
        widgets=[progressbar.Bar("=", "[", "]"), " ", progressbar.Percentage()],
    )
    bar.start()

    for idx, file in enumerate(files):
        results = ocr_png.get_value(f"results/{file}")
        for lang, data in results.items():
            data["No"] = idx + 1
            if data["File_Pdf"] == "":
                data["File_Pdf"] = file.split(".png")[0]
            if lang not in dict_of_dfs.keys():
                dict_of_dfs[lang] = pd.DataFrame()
            dict_of_dfs[lang] = dict_of_dfs[lang].append(data, ignore_index=True)

        bar.update(idx + 1)
    bar.finish()

    return dict_of_dfs


def write_to_excel(dict_of_dfs, excel_name):
    output_excel = excel_name

    writer = pd.ExcelWriter(output_excel)

    for lang, df in dict_of_dfs.items():
        for col in df.keys():
            if col == "Bruto" or col == "DPP" or col == "PPh_21" or col == "No":
                df[col] = df[col].astype(int)

        df = df[
            [
                "No",
                "File_Pdf",
                "Name",
                "NPWP",
                "No_Bukpot",
                "Date",
                "JENIS PAJAK",
                "Bruto",
                "DPP",
                "PPh_21",
            ]
        ]
        df.to_excel(writer, sheet_name=lang, index=False)

    writer.save()
    return output_excel


def add_link_color_comma(file_name):
    wb = load_workbook(file_name)

    ws = wb.active

    c_int_letter = []
    c_str_letter = []

    red_color = "FF0000"

    highlight = NamedStyle(name="highlight")
    highlight.font = Font(color=red_color)
    bd = Side(style="thin", color=red_color)
    highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    for header in ws[1]:
        if header.value == "Bruto" or header.value == "DPP" or header.value == "PPh_21":
            c_int_letter.append(header.column_letter)

        elif header.value == "File_Pdf":
            c_link_letter = header.column_letter

        elif header.value == "NPWP":
            c_str_letter.append(header.column_letter)
            l_npwp = header.column_letter

        elif header.value == "Date":
            c_str_letter.append(header.column_letter)

        elif header.value == "No_Bukpot":
            l_bukpot = header.column_letter

    list_sheet_kosong = []

    for sheet in wb.worksheets:
        for sheet2 in wb.worksheets:
            if sheet is not sheet2 and f"{sheet}{sheet2}" not in list_sheet_kosong:
                compare(sheet, sheet2)
                list_sheet_kosong.append(f"{sheet}{sheet2}")
                list_sheet_kosong.append(f"{sheet2}{sheet}")

        row_count = sheet.max_row

        sheet.conditional_formatting.add(
            f"{l_bukpot}1:{l_bukpot}{row_count}",
            FormulaRule(
                formula=[
                    f"COUNTIFS(${l_bukpot}$1:${l_bukpot}${row_count},{l_bukpot}1,${l_npwp}$1:${l_npwp}${row_count},{l_npwp}1)>1"
                ],
                font=highlight.font,
                border=highlight.border,
            ),
        )

        for idx in range(2, row_count + 1):
            filename = sheet[f"{c_link_letter}{idx}"].value

            sheet[f"{c_link_letter}{idx}"].hyperlink = f"{folder_png}/{filename}.png"
            sheet[f"{c_link_letter}{idx}"].font = Font(color="000000FF")

            for str_letter in c_str_letter:
                sheet[f"{str_letter}{idx}"].number_format = "@"

            for int_letter in c_int_letter:
                sheet[f"{int_letter}{idx}"].number_format = "#,##0"

    wb.save(file_name)


def compare(sheet1, sheet2):
    print(f"membandingkan {sheet1} dan {sheet2}")

    first_l = sheet1[1][0].column_letter
    last_l = sheet1[1][sheet1.max_column - 1].column_letter
    last_row = sheet1.max_row

    greenFill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    sheet1.conditional_formatting.add(
        f"{first_l}1:{last_l}{last_row}",
        CellIsRule(
            operator="notEqual", formula=[f"{sheet2.title}!{first_l}1"], fill=greenFill,
        ),
    )

    sheet2.conditional_formatting.add(
        f"{first_l}1:{last_l}{last_row}",
        CellIsRule(
            operator="notEqual", formula=[f"{sheet1.title}!{first_l}1"], fill=greenFill,
        ),
    )


folder_png = "results"
excel_name = "Incubator_Bukpot_Garry.xlsx"

files = make_list_of_files(folder_png)
dict_of_dfs = make_dict_of_dfs(files)
write_to_excel(dict_of_dfs, excel_name)
add_link_color_comma(excel_name)

print(f"\nDone\n\nFile Excel name: {excel_name}\n")
